import os
import sys
import csv
import time
import threading
import openpyxl
from datetime import datetime
import pytz
from openpyxl.utils import get_column_letter
from colorama import init, Fore
from dotenv import load_dotenv

init(autoreset=True)

class Storage:
    def __init__(self):
        self._load_env_file()
        # 创建数据存储目录
        self.data_dir = os.path.join(os.getcwd(), 'data')
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            print(Fore.GREEN + f"创建数据存储目录: {self.data_dir}")
        
        # 设置文件路径（每次实例化时生成新的时间戳）
        self.update_file_paths()
        
        # 记录是否已经写入过表头
        self.csv_header_written = False
        
        # 添加线程锁，防止多线程写入冲突
        self.write_lock = threading.Lock()

    def _load_env_file(self):
        """加载.env文件，支持打包后的程序"""
        try:
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
                env_file = os.path.join(base_path, '.env')
                if not os.path.exists(env_file):
                    resources_path = os.path.join(os.path.dirname(base_path), 'Resources')
                    env_file = os.path.join(resources_path, '.env')
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
                env_file = os.path.join(base_path, '.env')
            
            if os.path.exists(env_file):
                load_dotenv(env_file)
                print(Fore.GREEN + f"已加载环境配置文件: {env_file}")
            else:
                print(Fore.YELLOW + f"未找到环境配置文件: {env_file}")
        except Exception as e:
            print(Fore.RED + f"加载环境配置文件失败: {str(e)}")
    
    def append_single_to_csv(self, data_item):
        """实时追加单条数据到CSV文件"""
        if not data_item:
            print(Fore.RED + "没有数据需要追加")
            return False
        
        # 获取绝对路径
        output_path = os.path.abspath(self.csv_file)
        
        # 使用线程锁确保文件写入的原子性
        with self.write_lock:
            try:
                # 检查文件是否存在
                file_exists = os.path.exists(self.csv_file)
                headers = []
                
                if file_exists:
                    # 如果文件存在，读取现有表头
                    with open(self.csv_file, 'r', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile)
                        headers = next(reader, [])
                else:
                    # 如果文件不存在，使用新数据的表头
                    headers = list(data_item.keys())
                    # 将用户手机号字段移动到第一列
                    mobile_field = 'Other Information_User Mobile No'
                    if mobile_field in headers:
                        headers.remove(mobile_field)
                        headers.insert(0, mobile_field)
                
                # 检查是否有新字段
                new_fields = [key for key in data_item.keys() if key not in headers]
                
                if new_fields:
                    print(Fore.YELLOW + f"发现新字段: {new_fields}，更新表头")
                    # 添加新字段到表头
                    headers.extend(new_fields)
                    # 如果文件已存在，重新写入文件
                    if file_exists:
                        with open(self.csv_file, 'r', encoding='utf-8') as csvfile:
                            reader = csv.reader(csvfile)
                            existing_headers = next(reader, [])
                            data_rows = list(reader)
                        
                        with open(self.csv_file, 'w', newline='', encoding='utf-8') as csvfile:
                            writer = csv.writer(csvfile)
                            writer.writerow(headers)
                            # 重新写入现有数据
                            for row in data_rows:
                                # 确保每行数据长度与表头一致
                                while len(row) < len(headers):
                                    row.append('')
                                writer.writerow(row[:len(headers)])
                
                # 处理复杂数据类型并准备写入数据
                row_data = []
                for header in headers:
                    value = data_item.get(header, '')
                    if isinstance(value, (dict, list)):
                        value = str(value)
                    # 格式化时间字段（处理所有包含时间的字段）
                    elif 'Time' in header and value:
                        try:
                            # 检查是否为毫秒时间戳
                            if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                                timestamp = int(value)
                                # 如果是毫秒时间戳，转换为秒
                                if timestamp > 10000000000:
                                    timestamp = timestamp / 1000
                                # 转换为西非时间（UTC+1）
                                wat_tz = pytz.timezone('Africa/Lagos')
                                dt = datetime.fromtimestamp(timestamp, wat_tz)
                                value = dt.strftime('%Y-%m-%d %H:%M:%S')
                        except Exception as e:
                            print(Fore.YELLOW + f"时间格式化失败: {str(e)}")
                    row_data.append(value)
                
                # 写入CSV文件
                with open(self.csv_file, 'a', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    # 如果文件不存在，先写入表头
                    if not file_exists:
                        writer.writerow(headers)
                    # 写入数据行
                    writer.writerow(row_data)
                    csvfile.flush()  # 立即刷新缓冲区，确保数据写入磁盘
                
                print(Fore.GREEN + f"数据已追加到CSV文件: {output_path}")
                return True
                
            except Exception as e:
                print(Fore.RED + f"追加CSV文件失败: {str(e)}")
                # 即使失败，也返回True，继续处理下一条数据
                return True
    
    def update_file_paths(self):
        """更新文件路径，生成带有时间戳的文件名"""
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        self.excel_file = os.path.join(self.data_dir, f'{timestamp}_order_details.xlsx')
        self.csv_file = os.path.join(self.data_dir, f'{timestamp}_order_details.csv')
    
    def save_to_excel(self, data_list):
        """将数据保存到Excel"""
        if not data_list:
            print(Fore.RED + "没有数据需要保存")
            return False
        
        # 获取绝对路径
        output_path = os.path.abspath(self.excel_file)
        print(Fore.CYAN + f"开始保存数据到Excel文件: {output_path}")
        print(Fore.YELLOW + f"文件将保存在: {os.path.dirname(output_path)}")
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "订单详情"
            
            # 获取表头
            if data_list:
                headers = list(data_list[0].keys())
                
                # 写入表头
                for col, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col)
                    ws[f'{col_letter}1'] = header
                
                # 写入数据
                for row, item in enumerate(data_list, 2):
                    for col, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col)
                        value = item.get(header, '')
                        # 处理复杂数据类型
                        if isinstance(value, (dict, list)):
                            value = str(value)
                        # 格式化时间字段
                        elif header in ['settlement_time', 'create_time'] and value:
                            try:
                                # 检查是否为毫秒时间戳
                                if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                                    timestamp = int(value)
                                    # 如果是毫秒时间戳，转换为秒
                                    if timestamp > 10000000000:
                                        timestamp = timestamp / 1000
                                    # 转换为西非时间（UTC+1）
                                    wat_tz = pytz.timezone('Africa/Lagos')
                                    dt = datetime.fromtimestamp(timestamp, wat_tz)
                                    value = dt.strftime('%Y-%m-%d %H:%M:%S')
                            except Exception as e:
                                print(Fore.YELLOW + f"时间格式化失败: {str(e)}")
                        ws[f'{col_letter}{row}'] = value
                
                # 自动调整列宽
                for col in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col)
                    max_length = 0
                    for cell in ws[col_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
                    ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件
            wb.save(self.excel_file)
            print(Fore.GREEN + f"数据保存成功，文件路径: {output_path}")
            print(Fore.GREEN + f"文件位置: {os.path.dirname(output_path)}")
            print(Fore.GREEN + f"共保存 {len(data_list)} 条数据")
            return True
            
        except Exception as e:
            print(Fore.RED + f"保存Excel文件失败: {str(e)}")
            return False
    
    def append_to_excel(self, data_list):
        """追加数据到现有Excel文件"""
        if not data_list:
            print(Fore.RED + "没有数据需要追加")
            return False
        
        # 获取绝对路径
        output_path = os.path.abspath(self.excel_file)
        print(Fore.CYAN + f"开始追加数据到Excel文件: {output_path}")
        print(Fore.YELLOW + f"文件位置: {os.path.dirname(output_path)}")
        
        try:
            if os.path.exists(self.excel_file):
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "订单详情"
                start_row = 1
            
            # 获取表头
            if data_list:
                headers = list(data_list[0].keys())
                
                # 如果是新文件，写入表头
                if start_row == 1:
                    for col, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col)
                        ws[f'{col_letter}1'] = header
                    start_row = 2
                
                # 写入数据
                for row, item in enumerate(data_list, start_row):
                    for col, header in enumerate(headers, 1):
                        col_letter = get_column_letter(col)
                        value = item.get(header, '')
                        if isinstance(value, (dict, list)):
                            value = str(value)
                        # 格式化时间字段
                        elif header in ['settlement_time', 'create_time'] and value:
                            try:
                                # 检查是否为毫秒时间戳
                                if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                                    timestamp = int(value)
                                    # 如果是毫秒时间戳，转换为秒
                                    if timestamp > 10000000000:
                                        timestamp = timestamp / 1000
                                    # 转换为西非时间（UTC+1）
                                    wat_tz = pytz.timezone('Africa/Lagos')
                                    dt = datetime.fromtimestamp(timestamp, wat_tz)
                                    value = dt.strftime('%Y-%m-%d %H:%M:%S')
                            except Exception as e:
                                print(Fore.YELLOW + f"时间格式化失败: {str(e)}")
                        ws[f'{col_letter}{row}'] = value
                
                # 自动调整列宽
                for col in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col)
                    max_length = 0
                    for cell in ws[col_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件
            wb.save(self.excel_file)
            print(Fore.GREEN + f"数据追加成功，文件路径: {output_path}")
            print(Fore.GREEN + f"文件位置: {os.path.dirname(output_path)}")
            print(Fore.GREEN + f"共追加 {len(data_list)} 条数据")
            return True
            
        except Exception as e:
            print(Fore.RED + f"追加Excel文件失败: {str(e)}")
            return False
    
    def save_to_csv(self, data_list):
        """将数据保存到CSV"""
        if not data_list:
            print(Fore.RED + "没有数据需要保存")
            return False
        
        # 获取绝对路径
        output_path = os.path.abspath(self.csv_file)
        print(Fore.CYAN + f"开始保存数据到CSV文件: {output_path}")
        print(Fore.YELLOW + f"文件将保存在: {os.path.dirname(output_path)}")
        
        try:
            # 获取表头
            if data_list:
                headers = list(data_list[0].keys())
                
                # 将用户手机号字段移动到第一列
                mobile_field = 'Other Information_User Mobile No'
                if mobile_field in headers:
                    headers.remove(mobile_field)
                    headers.insert(0, mobile_field)
                
                # 写入CSV文件
                with open(self.csv_file, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=headers)
                    writer.writeheader()
                    
                    for item in data_list:
                        # 处理复杂数据类型
                        row = {}
                        for key, value in item.items():
                            if isinstance(value, (dict, list)):
                                value = str(value)
                            # 格式化时间字段
                            elif key in ['settlement_time', 'create_time'] and value:
                                try:
                                    # 检查是否为毫秒时间戳
                                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                                        timestamp = int(value)
                                        # 如果是毫秒时间戳，转换为秒
                                        if timestamp > 10000000000:
                                            timestamp = timestamp / 1000
                                        # 转换为西非时间（UTC+1）
                                        wat_tz = pytz.timezone('Africa/Lagos')
                                        dt = datetime.fromtimestamp(timestamp, wat_tz)
                                        value = dt.strftime('%Y-%m-%d %H:%M:%S')
                                except Exception as e:
                                    print(Fore.YELLOW + f"时间格式化失败: {str(e)}")
                            row[key] = value
                        writer.writerow(row)
            
            print(Fore.GREEN + f"数据保存成功，文件路径: {output_path}")
            print(Fore.GREEN + f"文件位置: {os.path.dirname(output_path)}")
            print(Fore.GREEN + f"共保存 {len(data_list)} 条数据")
            return True
            
        except Exception as e:
            print(Fore.RED + f"保存CSV文件失败: {str(e)}")
            return False
    
    def append_to_csv(self, data_list):
        """追加数据到现有CSV文件"""
        if not data_list:
            print(Fore.RED + "没有数据需要追加")
            return False
        
        # 获取绝对路径
        output_path = os.path.abspath(self.csv_file)
        print(Fore.CYAN + f"开始追加数据到CSV文件: {output_path}")
        print(Fore.YELLOW + f"文件位置: {os.path.dirname(output_path)}")
        
        try:
            # 检查文件是否存在
            file_exists = os.path.exists(self.csv_file)
            
            # 获取表头
            if data_list:
                headers = list(data_list[0].keys())
                
                # 将用户手机号字段移动到第一列（仅当创建新文件时）
                if not file_exists:
                    mobile_field = 'Other Information_User Mobile No'
                    if mobile_field in headers:
                        headers.remove(mobile_field)
                        headers.insert(0, mobile_field)
                
                # 写入CSV文件
                with open(self.csv_file, 'a' if file_exists else 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=headers)
                    
                    # 如果是新文件，写入表头
                    if not file_exists:
                        writer.writeheader()
                    
                    for item in data_list:
                        # 处理复杂数据类型
                        row = {}
                        for key, value in item.items():
                            if isinstance(value, (dict, list)):
                                value = str(value)
                            # 格式化时间字段
                            elif key in ['settlement_time', 'create_time'] and value:
                                try:
                                    # 检查是否为毫秒时间戳
                                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                                        timestamp = int(value)
                                        # 如果是毫秒时间戳，转换为秒
                                        if timestamp > 10000000000:
                                            timestamp = timestamp / 1000
                                        # 转换为西非时间（UTC+1）
                                        wat_tz = pytz.timezone('Africa/Lagos')
                                        dt = datetime.fromtimestamp(timestamp, wat_tz)
                                        value = dt.strftime('%Y-%m-%d %H:%M:%S')
                                except Exception as e:
                                    print(Fore.YELLOW + f"时间格式化失败: {str(e)}")
                            row[key] = value
                        writer.writerow(row)
            
            print(Fore.GREEN + f"数据追加成功，文件路径: {output_path}")
            print(Fore.GREEN + f"文件位置: {os.path.dirname(output_path)}")
            print(Fore.GREEN + f"共追加 {len(data_list)} 条数据")
            return True
            
        except Exception as e:
            print(Fore.RED + f"追加CSV文件失败: {str(e)}")
            return False
    
    def get_data_dir(self):
        """获取数据存储目录"""
        return self.data_dir
